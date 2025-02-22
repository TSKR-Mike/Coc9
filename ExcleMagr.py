"""
load_workbook(): 用于加载Excel文档。
active cell: 获取或选择工作表的活动单元格。
get_cell_value(): 获取单元格的值。
set_cell_value(value): 设置单元格的值。
cell(row=None, column=None): 获取或选择工作表的单元格，其中row参数指定行号，column参数指定列号。
column(index=None): 返回工作表中指定列的索引，如果不存在则返回默认值0。
row(): 返回工作表中指定行的索引，如果不存在则返回默认值0。
get_column_letter(column_index=None): 返回指定列的字母缩写，如果不存在则返回默认值'A'。
get_rows_count(): 返回工作表中行的数量。
get_worksheet_name(): 返回工作表的名称。
get_sheet_by_name(name=None): 根据名称获取工作表对象，如果不存在则返回默认值None。
number_format(column=None, row=None, guarantee_legal=True): 设置单元格的格式，包括数字格式、货币格式、日期格式等。
apply_format(cell, format): 应用单元格格式。
merge_cells(start_row=1, start_column=1, end_row=None, end_column=None, merge_type='left'): 合并单元格，可以指定合并方式（'left'，'right'，'inner'）。
cell_value_is_string(): 检查单元格是否为字符串。
cell_value_is_number(): 检查单元格是否为数字。
cell_value_is_datetime(): 检查单元格是否为日期或时间。
cell_value_is_boolean(): 检查单元格是否为布尔值。
convert_chars(from_char, to_char, from_unit, to_unit): 将单元格中的字符转换为另一种字符类型。
upper(): 将单元格中的所有字符转换为大写字母。
lower(): 将单元格中的所有字符转换为小写字母。
title(): 将单元格中的文本标题转换为大写字母。
left(width): 向左填充指定宽度的空格。
right(width): 向右填充指定宽度的空格。
fill(width): 填充指定宽度的空格。
justify(width): 在指定方向上对单元格中的内容进行对齐。
wrap(width): 将单元格中的内容进行换行。
"""
# user guide

import openpyxl


class ExcelMgr:
    def __init__(self, file_name):

        self.filename = file_name
        self.data = []
        self.f_close = True
        self.read_excel_data()

    def read_excel_data(self):
        """
        该函数返回excel文件的所有行和列
        """
        # 打开Excel文件
        try:
            workbook = openpyxl.load_workbook(self.filename, keep_vba=True)
        except PermissionError:
            self.sheet = self.data = None
            raise PermissionError(
                "The file '" + self.filename + "' was locked by other programs,please close the program and then try "
                                               "again.")
        except FileNotFoundError:
            self.sheet = self.data = None
            raise FileNotFoundError("The xlsx file '" + self.filename + "' is not exist or the dictionary is false.")
            # 选择第一个工作表
        self.f_close = False
        self.workbook = workbook
        sheet = workbook.active
        self.sheet = sheet
        data = []
        # 遍历行和列，读取数据

        for row in sheet.iter_rows():
            k = []
            for cell in row:
                k.append(cell.value)
                #  print(cell.value)
            data.append(k)
        # print(data)
        self.data = data

    def get_cell_value(self, x, y):
        if self.sheet is not None:
            if type(x) != int or type(y) != int:
                raise TypeError('not input a int number')
            try:
                return self.data[x][y]
            except IndexError:
                raise IndexError(
                    'INDEX ERROR：out of range(x=' + str(x) + ',max x =' + str(len(self.data)) + ', y=' + str(
                        y) + 'max y =' + str(len(self.data[0])))
        else:
            return -1

    def read_lines(self, line_num):
        if self.sheet is not None:
            return [i for i in self.data[line_num]]
        else:
            return -1

    def read_y_line(self, num):
        if self.sheet is not None:
            return [i[num] for i in self.data]
        else:
            return -1

    # these are the movements that can change the cells` data
    def write_cell_value(self, x, y, value):
        if self.sheet is not None:
            if type(x) != int or type(y) != int or (type(value) != int and type(value) != float and type(value) != str):
                raise TypeError
            try:
                self.sheet.cell(x, y).value = value
                self.read_excel_data()
            except IndexError:
                raise IndexError(
                    'INDEX ERROR：out of range(x=' + str(x) + ',max x =' + str(len(self.data)) + ', y=' + str(
                        y) + 'max y =' + str(len(self.data[0])))
        else:
            return -1

    def write_x_lines(self, line_num, values):
        if self.sheet is not None:
            for i in range(len(values)):
                self.sheet.cell(line_num, i + 1).value = values[i]
            self.read_excel_data()  # update the data after writing
        else:
            return -1

    def write_y_line(self, num, values):
        if self.sheet is not None:
            for i in range(len(self.data[num])):
                self.sheet.cell(i + 1, num).value = values[i]
            self.read_excel_data()  # update the data after writing
        else:
            return -1
    def save(self):
        if self.sheet is not None:
            try:
                self.workbook.save(self.filename)
                self.workbook.close()
            except PermissionError:
                raise PermissionError(
                    "The file '" + self.filename + "' was locked by other programs,please close the program and then try again.")

        else:
            return -1

    def __delete__(self, instance):
        if not self.f_close:
            self.save()

if __name__ == '__main__':
    mgr = ExcelMgr('06.xlsx')
    mgr.read_excel_data()
    print(mgr.data)
